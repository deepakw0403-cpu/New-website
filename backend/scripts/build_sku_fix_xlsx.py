"""
One-off script that audits every SKU and writes an Excel workbook with:
- Column A : Current Name
- Column B : Seller
- Column C : Category
- Column D : Issues detected (semicolon-separated)
- Column E : Suggested New Name
- Column F : Other Field Fixes (HSN, MOQ, width, composition structure, etc.)
- Column G : SKU ID (short, for admin search)

Usage:
    cd /app/backend && python3 scripts/build_sku_fix_xlsx.py
Output:
    /app/memory/sku_fix_suggestions.xlsx
"""
import os
import re
import asyncio
import sys
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


CANONICAL = {
    "Cotton", "Organic Cotton", "Recycled Cotton", "Polyester",
    "Recycled Polyester", "Viscose", "Lyocell", "Modal", "Lycra", "Linen",
    "Hemp", "Nylon", "Wool", "Silk", "Bamboo", "Acrylic", "Cashmere",
    "Lurex", "Jute", "Rayon", "Spandex", "Elastane", "Tencel", "Cupro", "Ramie",
}

MATERIAL_ALIASES = {
    "poly": "Polyester", "polyester": "Polyester",
    "cotton": "Cotton", "coitton": "Cotton", "cottton": "Cotton",
    "org cotton": "Organic Cotton", "org. cotton": "Organic Cotton",
    "organic cotton": "Organic Cotton", "recycled cotton": "Recycled Cotton",
    "recycled polyester": "Recycled Polyester",
    "lycra": "Lycra", "lyvra": "Lycra", "spandex": "Spandex",
    "elastane": "Elastane", "hemp": "Hemp", "viscose": "Viscose",
    "linen": "Linen", "tencel": "Tencel", "lyocell": "Lyocell",
    "modal": "Modal", "nylon": "Nylon", "wool": "Wool", "silk": "Silk",
    "bamboo": "Bamboo", "acrylic": "Acrylic", "rayon": "Rayon",
}

WEAVE_PATTERNS = [
    (r"\b3\s*[\\/]\s*1\s*RHT\b", "3/1 RHT"),
    (r"\b3\s*[\\/]\s*1\s*LHT\b", "3/1 LHT"),
    (r"\b2\s*[\\/]\s*1\s*RHT\b", "2/1 RHT"),
    (r"\b2\s*[\\/]\s*1\s*LHT\b", "2/1 LHT"),
    (r"\b3\s*[\\/]\s*1\s*Twill\b", "3/1 Twill"),
    (r"\b2\s*[\\/]\s*2\s*Twill\b", "2/2 Twill"),
    (r"\b4\s*[\\/]\s*1\s*Satin\b", "4/1 Satin"),
    (r"\bHerringbone\b", "Herringbone"),
    (r"\bDobby\b", "Dobby"),
]

HSN_BY_CATEGORY = {
    "Denim": "5209",
    "Cotton Fabrics": "5208",
    "Polyester Fabrics": "5407",
    "Linen": "5309",
    "Sustainable Fabrics": "5311",
    "Viscose": "5408",
    "Cotton Knits": "6006",
}


def normalise_material(raw: str) -> str:
    key = (raw or "").strip().lower()
    return MATERIAL_ALIASES.get(key, (raw or "").strip().title())


def detect_weave(text: str) -> str:
    for rx, label in WEAVE_PATTERNS:
        if re.search(rx, text, re.IGNORECASE):
            return label
    return ""


def build_materials_str(fabric: dict) -> str:
    comp = fabric.get("composition")
    if isinstance(comp, list):
        mats = []
        for c in comp:
            if not isinstance(c, dict):
                continue
            m = normalise_material(c.get("material", ""))
            if m and m not in mats:
                mats.append(m)
            if len(mats) == 3:
                break
        return " ".join(mats)
    # Composition is a string → extract material names heuristically
    if isinstance(comp, str):
        found = []
        for token in re.split(r"[,+/&]", comp):
            low = token.strip().lower()
            low = re.sub(r"^\d+\s*%\s*", "", low).strip()
            m = normalise_material(low)
            if m in CANONICAL and m not in found:
                found.append(m)
            if len(found) == 3:
                break
        return " ".join(found)
    return ""


def clean_ounce(ounce) -> str:
    if ounce is None or ounce == "":
        return ""
    s = str(ounce)
    m = re.match(r"^\s*([\d.]+)", s)
    if not m:
        return ""
    num = m.group(1).rstrip(".")
    if "." in num:
        num = num.rstrip("0").rstrip(".")
    return num


def suggest_new_name(fabric: dict, cat_name: str) -> str:
    """Return the name that this SKU *should* have per the format spec."""
    mats = build_materials_str(fabric) or ""
    existing_name = fabric.get("name", "")

    # Detect weave
    weave = (fabric.get("weave_type") or "").strip()
    if not weave:
        weave = detect_weave(existing_name)

    # Construction or AOP/pattern hints from the existing name
    construction = ""
    # Try to extract non-material, non-weave, non-weight words from existing name
    KNOWN_CONSTRUCTIONS = [
        "Poplin", "Twill", "Jersey", "Interlock", "Rib", "Fleece", "Taffeta",
        "Satin", "Seersucker", "Slub", "Ripstop", "Voile", "Canvas",
        "Oxford", "Chambray", "Sateen", "Single Jersey", "Double Cloth",
        "Yarn Dyed", "Dobby AOP", "AOP", "Jacquard", "Paper Poplin",
        "Looper Jacquard", "Dot Knit", "Rice Knit", "Micro Crepe",
    ]
    for k in KNOWN_CONSTRUCTIONS:
        if re.search(rf"\b{re.escape(k)}\b", existing_name, re.IGNORECASE):
            construction = k if k != "Dobby AOP" else "Dobby AOP"
            break

    # Weight
    weight = ""
    weight_unit = fabric.get("weight_unit", "")
    if weight_unit == "ounce" and fabric.get("ounce"):
        oz = clean_ounce(fabric["ounce"])
        if oz:
            weight = f"{oz}oz"
    elif fabric.get("gsm"):
        weight = f"{fabric['gsm']} GSM"
    elif fabric.get("ounce"):
        oz = clean_ounce(fabric["ounce"])
        if oz:
            weight = f"{oz}oz"

    is_denim = cat_name == "Denim"

    if is_denim:
        # Denim: "<mats>, <weave>, <weight>, Color: <color>"
        color = (fabric.get("color") or "").strip()
        has_multi = fabric.get("has_multiple_colors")
        parts = [p for p in [mats, weave, weight] if p]
        if not parts:
            return ""  # cannot infer
        base = ", ".join(parts)
        if has_multi:
            # Keep original "Color: X x Y" if already in name
            m = re.search(r"Color:\s*[^,]+", existing_name, re.IGNORECASE)
            if m:
                return f"{base}, {m.group(0)}"
            return base
        if color:
            return f"{base}, Color: {color}"
        return base
    else:
        # Non-denim: "<mats> <construction>, <GSM> GSM"
        # Backfill construction from remaining name words if not found
        if not construction:
            leftover = existing_name
            # Strip materials
            for tok in (mats or "").split():
                leftover = re.sub(rf"\b{re.escape(tok)}\b", "", leftover, flags=re.IGNORECASE)
            # Strip weight
            leftover = re.sub(r"\b\d+(?:\.\d+)?\s*(?:GSM|oz|OZ)\b", "", leftover)
            # Strip weave
            leftover = re.sub(r"\b\d\s*[\\/]\s*\d\s*\w+\b", "", leftover)
            # Strip apostrophe-count noise
            leftover = re.sub(r"\d+'s\s*[xX]\s*\d+'s?", "", leftover)
            leftover = re.sub(r"\s{2,}", " ", leftover).strip(" ,")
            # Title-case remainder
            construction = leftover.strip().title() if leftover.strip() else ""
        # Normalise backslashes in construction to slashes
        construction = construction.replace("\\", "/").strip()

        left_part = " ".join(p for p in [mats, construction] if p).strip()
        # Clean residual double-spaces
        left_part = re.sub(r"\s{2,}", " ", left_part).strip()
        # Replace stray Poly → Polyester, typos
        for k, v in MATERIAL_ALIASES.items():
            if k == "poly":
                left_part = re.sub(r"\bPoly\b", "Polyester", left_part)
            elif k == "cottton":
                left_part = re.sub(r"\bCottton\b", "Cotton", left_part, flags=re.IGNORECASE)

        if weight:
            return f"{left_part}, {weight}" if left_part else weight
        return left_part


def detect_issues(fabric: dict, cat_name: str, dup_group_ids: list) -> list:
    issues = []
    name = fabric.get("name", "")
    comp = fabric.get("composition")

    # Composition structure
    if isinstance(comp, str):
        issues.append("Composition stored as string (not array) — must re-enter via Admin composition dropdown")
    elif not comp:
        issues.append("Composition empty")
    elif isinstance(comp, list):
        bad = [c.get("material", "") for c in comp if isinstance(c, dict) and c.get("material", "") not in CANONICAL]
        if bad:
            issues.append(f"Non-canonical material(s): {bad}")
        total = sum(float(c.get("percentage") or 0) for c in comp if isinstance(c, dict))
        if 0 < total < 99.99 or total > 100.01:
            issues.append(f"Composition percentages sum to {total}% (should be 100%)")

    # Name format
    if "\\" in name:
        issues.append("Backslash in weave notation (use `/`, not `\\`)")
    if re.search(r"\d+'s", name):
        issues.append("Apostrophe in yarn count (`40's x 40's` → `40x40`)")
    if "  " in name:
        issues.append("Double-space in name")
    if re.search(r"\bPoly\b", name):
        issues.append("`Poly` → `Polyester`")
    if re.search(r"\bCottton\b", name, re.IGNORECASE):
        issues.append("Typo: `Cottton` → `Cotton`")
    if re.search(r"\bLyvra\b", name, re.IGNORECASE):
        issues.append("Typo: `Lyvra` → `Lycra`")
    words = name.split()
    if words and words[0] and words[0][0].islower():
        issues.append("First word is lowercase — use Title Case")

    # Weight in name
    if not re.search(r"\d+\s*(GSM|oz|OZ)", name):
        issues.append("Weight (GSM/oz) missing from name")

    # Denim-specific
    if cat_name == "Denim":
        if not fabric.get("weave_type"):
            issues.append("`weave_type` field empty")
        if not fabric.get("ounce"):
            issues.append("`ounce` field empty")
        if "color:" not in name.lower() and not fabric.get("has_multiple_colors"):
            issues.append("Color missing from denim name")

    # Other required fields
    if not fabric.get("width"):
        issues.append("Width missing")
    if not fabric.get("moq"):
        issues.append("MOQ missing")
    if not (fabric.get("images") or []) and not any(
        (isinstance(v, dict) and v.get("image_url")) for v in (fabric.get("color_variants") or [])
    ):
        issues.append("Images missing")
    if not fabric.get("hsn_code"):
        suggested = HSN_BY_CATEGORY.get(cat_name, "<confirm with GST consultant>")
        issues.append(f"HSN code missing (suggested: {suggested})")
    if not fabric.get("fabric_code"):
        issues.append("`fabric_code` missing")

    # Bookable-but-no-price
    if fabric.get("is_bookable"):
        if not (fabric.get("rate_per_meter") or 0) > 0:
            issues.append("Marked bookable but `rate_per_meter` is empty")
        if not (fabric.get("quantity_available") or 0) > 0:
            issues.append("Marked bookable but `quantity_available` is 0")

    # Duplicate
    if len(dup_group_ids) > 1:
        others = [i[:8] for i in dup_group_ids if i != fabric.get("id")]
        issues.append(f"Duplicate name with {len(others)} other SKU(s) — differentiate by GSM/weight or delete ({', '.join(others)})")

    return issues


def field_fixes(fabric: dict, cat_name: str) -> list:
    """Non-name field fixes that must be entered in the Admin form."""
    fixes = []
    if not fabric.get("hsn_code"):
        suggested = HSN_BY_CATEGORY.get(cat_name, "")
        if suggested:
            fixes.append(f"Set HSN = {suggested}")
    if not fabric.get("moq"):
        if cat_name == "Cotton Knits" or fabric.get("fabric_type") == "knitted":
            fixes.append("Set MOQ = 150 KG")
        elif cat_name == "Denim":
            fixes.append("Set MOQ = 500 MTR")
        else:
            fixes.append("Set MOQ = 1500 MTR")
    if not fabric.get("width"):
        fixes.append("Set width (commonly 58\" or 54\" — confirm with mill)")
    if isinstance(fabric.get("composition"), str):
        fixes.append(f"Composition: re-enter via Admin dropdown (current raw: \"{fabric['composition']}\")")
    if cat_name == "Denim" and not fabric.get("weave_type"):
        fixes.append("Set `weave_type` from dropdown (3/1 RHT, 2/1 RHT, Dobby, etc.)")
    if cat_name == "Denim" and not fabric.get("ounce"):
        fixes.append("Set `ounce` (numeric)")
    if not fabric.get("fabric_code"):
        fixes.append("Generate `fabric_code`")
    return fixes


async def main():
    db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]
    cats = await db.categories.find({}, {"_id": 0}).to_list(500)
    cat_by_id = {c["id"]: c["name"] for c in cats}

    fabrics = await db.fabrics.find({}, {"_id": 0}).to_list(5000)

    # Build duplicate groups by (normalised name, seller)
    dup_map = defaultdict(list)
    for f in fabrics:
        norm = re.sub(r"\s+", " ", (f.get("name", "") or "").lower()).strip()
        key = (norm, f.get("seller_company", ""))
        dup_map[key].append(f.get("id", ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "SKU Fix Suggestions"

    headers = [
        "Current Name",
        "Seller",
        "Category",
        "Issues Detected",
        "Suggested New Name",
        "Other Field Fixes",
        "SKU ID (short)",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    p0_fill = PatternFill("solid", fgColor="FEE2E2")
    p1_fill = PatternFill("solid", fgColor="FEF3C7")

    rows_written = 0
    p0_kw = ("Composition stored as string", "Backslash", "`Poly`", "Typo", "lowercase", "weave_type` field empty", "Color missing from denim name")

    for f in sorted(fabrics, key=lambda x: (cat_by_id.get(x.get("category_id", ""), ""), x.get("name", "").lower())):
        cat_name = cat_by_id.get(f.get("category_id", ""), "Unknown")
        norm = re.sub(r"\s+", " ", (f.get("name", "") or "").lower()).strip()
        group = dup_map.get((norm, f.get("seller_company", "")), [])
        issues = detect_issues(f, cat_name, group)
        if not issues:
            continue  # Skip clean rows — data team only needs fix list
        suggested = suggest_new_name(f, cat_name)
        fixes = field_fixes(f, cat_name)

        row = [
            f.get("name", ""),
            f.get("seller_company", "") or "—",
            cat_name,
            "\n".join(f"• {i}" for i in issues),
            suggested or "(cannot infer — set manually)",
            "\n".join(f"• {x}" for x in fixes) if fixes else "",
            f.get("id", "")[:8],
        ]
        ws.append(row)
        rows_written += 1
        r = ws.max_row
        is_p0 = any(any(kw in i for kw in p0_kw) for i in issues)
        fill = p0_fill if is_p0 else p1_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).alignment = Alignment(vertical="top", wrap_text=True)
            if col == 4:  # Issues column
                ws.cell(row=r, column=col).fill = fill

    widths = [42, 24, 20, 55, 48, 38, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = Path("/app/memory/sku_fix_suggestions.xlsx")
    wb.save(out)
    print(f"\n✅ Wrote {rows_written} rows to {out}")
    print(f"   Total SKUs audited : {len(fabrics)}")
    print(f"   Clean rows skipped : {len(fabrics) - rows_written}")


if __name__ == "__main__":
    asyncio.run(main())
